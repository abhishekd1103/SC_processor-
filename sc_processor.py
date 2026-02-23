"""
╔══════════════════════════════════════════════════════════════════════════════╗
║         SHORT CIRCUIT STUDY PROCESSOR — Complete Single-File App            ║
║         IEC 60909 | LVCB / HVCB / BUS Auto-Classification                  ║
║         Dual Evaluation: I"k (Breaking) + ip (Peak)                         ║
║                                                                              ║
║  Deploy:  streamlit run sc_study_app.py                                      ║
║  GitHub:  Push this file + requirements.txt → Streamlit Cloud                ║
╚══════════════════════════════════════════════════════════════════════════════╝

requirements.txt contents:
    streamlit>=1.35.0
    pandas>=2.0.0
    openpyxl>=3.1.0
    XlsxWriter>=3.1.0
"""

# ══════════════════════════════════════════════════════════════════════════════
#  IMPORTS
# ══════════════════════════════════════════════════════════════════════════════
from __future__ import annotations

import io
import math
import warnings
from typing import Optional

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side,
)
from openpyxl.styles.numbers import FORMAT_NUMBER_00
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ══════════════════════════════════════════════════════════════════════════════
#  ❶  CONSTANTS & COLUMN MAPPING
# ══════════════════════════════════════════════════════════════════════════════

# ── Input column names (stripped of whitespace) ──────────────────────────────
C_ID            = "ID"               # Equipment identifier
C_KV            = "kV"               # Voltage level → drives LVCB/HVCB split
C_TYPE          = "Type"             # CB / Bus / etc.
C_CFACTOR       = "Cfactor"          # Voltage correction factor
C_BUS           = "Bus"              # Connected bus
C_RATED_IP      = "Rated ip"         # 🔴 RATED peak current (kA)
C_RATED_IB_SYM  = "Rated Ib Sym"    # 🔴 RATED breaking sym (kA) = Rated Ik
C_RATED_IB_ASYM = "Rated Ib asym"   # 🔴 RATED breaking asym (kA)
C_IK            = 'I"k'             # 🔵 SIMULATED initial SC current (kA)
C_IP            = "ip"               # 🔵 SIMULATED peak current (kA)
C_IB_SYM        = "Ib Sym"          # 🔵 Simulated breaking sym
C_IB_ASYM       = "Ib asym"         # 🔵 Simulated breaking asym
C_IDC           = "Idc"             # 🔵 DC component
C_ITH           = "Ith"             # 🔵 Thermal equivalent current
C_THERMAL       = "Thermal Energy"   # 🔵 Thermal energy
C_STANDARD      = "Standard"        # IEC standard applied

# ── Computed output columns ───────────────────────────────────────────────────
C_RATED_IK      = "Rated Ik (kA)"        # = Rated Ib Sym (explicit alias)
C_UTIL_IK       = 'Utilization % (I"k)'  # (I"k / Rated Ib Sym) × 100
C_UTIL_IP       = "Utilization % (ip)"   # (ip  / Rated ip)     × 100
C_RES_IK        = 'I"k Result'           # Pass / Fail per breaking check
C_RES_IP        = "ip Result"            # Pass / Fail / N/A per peak check
C_OVERALL       = "Overall Result"       # Combined verdict
C_REMARKS       = "Remarks"              # Engineer annotation placeholder

# ── Required minimum columns in the input ────────────────────────────────────
REQUIRED_COLS = [C_ID, C_KV, C_TYPE, C_RATED_IP, C_RATED_IB_SYM, C_IK, C_IP]

# ── Voltage threshold: ≤ 1 kV → LVCB, > 1 kV → HVCB ────────────────────────
LV_KV_THRESHOLD = 1.0

# ── Output sheet names ────────────────────────────────────────────────────────
SH_LVCB   = "LVCB"
SH_HVCB   = "HVCB"
SH_BUS    = "BUS"
SH_LEGEND = "Legend"

# ── Bus-type detection keywords ───────────────────────────────────────────────
BUS_KEYWORDS = ("bus", "busbar", "busduct")


# ══════════════════════════════════════════════════════════════════════════════
#  ❷  PURE EVALUATION LOGIC
# ══════════════════════════════════════════════════════════════════════════════

def _to_float(v: object) -> Optional[float]:
    """Safe coercion to float; returns None for blank / NaN / non-numeric."""
    if v is None:
        return None
    try:
        f = float(v)  # type: ignore[arg-type]
        return None if math.isnan(f) else f
    except (TypeError, ValueError):
        return None


def _ik_check(ik_sim: object, rated_ib_sym: object) -> str:
    """
    CHECK 1 — Breaking current.
    I"k (simulated) vs Rated Ib Sym (nameplate rated duty).
    """
    sim   = _to_float(ik_sim)
    rated = _to_float(rated_ib_sym)
    if sim is None or rated is None:
        return "N/A – Missing Data"
    return "Pass" if sim <= rated else "Fail"


def _ip_check(ip_sim: object, rated_ip: object) -> str:
    """
    CHECK 2 — Peak current.
    ip (simulated) vs Rated ip (nameplate peak withstand).
    Rated ip = 0 → equipment not rated for peak duty.
    """
    sim   = _to_float(ip_sim)
    rated = _to_float(rated_ip)
    if sim is None or rated is None:
        return "N/A – Missing Data"
    if rated == 0.0:
        return "N/A (Rated ip = 0)"
    return "Pass" if sim <= rated else "Fail"


def _util(sim: object, rated: object) -> Optional[float]:
    """Utilization % = (sim / rated) × 100. Returns None when undefined."""
    s = _to_float(sim)
    r = _to_float(rated)
    if s is None or r is None or r == 0.0:
        return None
    return round((s / r) * 100, 2)


def _overall(ik_res: str, ip_res: str) -> str:
    """Combine both sub-results into a single Overall Result verdict."""
    if "Fail" in (ik_res, ip_res):
        return "Fail"
    if ik_res == "Pass" and ip_res == "Pass":
        return "Pass"
    if ik_res == "Pass" and ip_res.startswith("N/A"):
        return "Pass (Ik only)"        # ip not rated — only Ik evaluated
    if ip_res == "Pass" and ik_res.startswith("N/A"):
        return "Pass (ip only)"
    return "N/A – Insufficient Data"


def _classify(kv: object, eq_type: object) -> str:
    """Return the output sheet name for one equipment row."""
    type_l = str(eq_type).strip().lower()
    if any(k in type_l for k in BUS_KEYWORDS):
        return SH_BUS
    try:
        return SH_LVCB if float(kv) <= LV_KV_THRESHOLD else SH_HVCB  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return SH_LVCB  # fallback


def evaluate_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Apply full evaluation to a cleaned DataFrame.

    Adds columns:
        Rated Ik (kA) | Util% Ik | Util% ip | I"k Result | ip Result |
        Overall Result | Remarks | _sheet (internal routing tag)
    """
    df = df.copy()

    df[C_RATED_IK] = df[C_RATED_IB_SYM]

    df[C_UTIL_IK] = df.apply(
        lambda r: _util(r[C_IK], r[C_RATED_IB_SYM]), axis=1
    )
    df[C_UTIL_IP] = df.apply(
        lambda r: _util(r[C_IP], r[C_RATED_IP]), axis=1
    )
    df[C_RES_IK] = df.apply(
        lambda r: _ik_check(r[C_IK], r[C_RATED_IB_SYM]), axis=1
    )
    df[C_RES_IP] = df.apply(
        lambda r: _ip_check(r[C_IP], r[C_RATED_IP]), axis=1
    )
    df[C_OVERALL] = df.apply(
        lambda r: _overall(r[C_RES_IK], r[C_RES_IP]), axis=1
    )
    df[C_REMARKS] = ""

    df["_sheet"] = df.apply(
        lambda r: _classify(r[C_KV], r[C_TYPE]), axis=1
    )
    return df


# ══════════════════════════════════════════════════════════════════════════════
#  ❸  EXCEL INPUT READER
# ══════════════════════════════════════════════════════════════════════════════

def read_and_clean(file_bytes: bytes) -> tuple[pd.DataFrame, list[str]]:
    """
    Read every sheet from the uploaded Excel, clean headers,
    validate columns, and concatenate into one DataFrame.

    Returns
    -------
    (combined_df, warnings_list)
    """
    xl = pd.ExcelFile(io.BytesIO(file_bytes), engine="openpyxl")
    frames: list[pd.DataFrame] = []
    warns: list[str] = []

    for sheet in xl.sheet_names:
        raw = xl.parse(sheet, keep_default_na=True)
        # Strip header whitespace (handles 'ID     ' → 'ID')
        raw.columns = [str(c).strip() for c in raw.columns]
        raw = raw.dropna(how="all").reset_index(drop=True)

        if raw.empty:
            warns.append(f"Sheet '{sheet}' is empty — skipped.")
            continue

        missing = [c for c in REQUIRED_COLS if c not in raw.columns]
        if missing:
            warns.append(
                f"Sheet '{sheet}' skipped — missing columns: {missing}"
            )
            continue

        raw["_source_sheet"] = sheet
        frames.append(raw)

    if not frames:
        raise ValueError(
            "No valid data sheets found. "
            f"Each sheet must contain: {REQUIRED_COLS}"
        )

    combined = pd.concat(frames, ignore_index=True)
    return combined, warns


# ══════════════════════════════════════════════════════════════════════════════
#  ❹  COLUMN ORDER FOR OUTPUT
# ══════════════════════════════════════════════════════════════════════════════

# Sections define order and background colour group in Excel
_INFO_COLS = [C_ID, C_KV, C_TYPE, C_CFACTOR, C_BUS]

_RATED_COLS_ORDER = [C_RATED_IP, C_RATED_IB_SYM, C_RATED_IB_ASYM, C_RATED_IK]

_SIM_COLS_ORDER = [C_IK, C_IP, C_IB_SYM, C_IB_ASYM, C_IDC, C_ITH, C_THERMAL]

_COMPUTED_COLS_ORDER = [
    C_UTIL_IK, C_UTIL_IP,
    C_RES_IK, C_RES_IP, C_OVERALL, C_REMARKS,
]

_TAIL_COLS = [C_STANDARD]

# Which section each column belongs to (for Excel colouring)
_COL_SECTION: dict[str, str] = {}
for _c in _INFO_COLS:        _COL_SECTION[_c] = "info"
for _c in _RATED_COLS_ORDER: _COL_SECTION[_c] = "rated"
for _c in _SIM_COLS_ORDER:   _COL_SECTION[_c] = "sim"
for _c in _COMPUTED_COLS_ORDER: _COL_SECTION[_c] = "computed"


def _ordered_cols(df: pd.DataFrame) -> list[str]:
    """Return final column order, keeping only columns that exist in df."""
    skip = {"_sheet", "_source_sheet"}
    explicit = (
        _INFO_COLS + _RATED_COLS_ORDER + _SIM_COLS_ORDER
        + _COMPUTED_COLS_ORDER + _TAIL_COLS
    )
    known = set(explicit) | skip
    extras = [c for c in df.columns if c not in known]
    ordered = explicit + extras
    return [c for c in ordered if c in df.columns and c not in skip]


# ══════════════════════════════════════════════════════════════════════════════
#  ❺  EXCEL OUTPUT FORMATTER
# ══════════════════════════════════════════════════════════════════════════════

# ── openpyxl colour/style helpers ────────────────────────────────────────────

def _fill(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex6)

def _font(color="000000", bold=False, size=9, italic=False) -> Font:
    return Font(name="Calibri", size=size, bold=bold, italic=italic,
                color=color)

def _border(thin="BFBFBF", thick="9DC3E6") -> Border:
    t = Side(style="thin",   color=thin)
    T = Side(style="medium", color=thick)
    return Border(left=t, right=t, top=t, bottom=t)

def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ── Per-sheet theme colours ───────────────────────────────────────────────────
SHEET_THEME = {
    SH_LVCB  : {"hdr": "1F4E79", "tab": "2E75B6", "hdr2": "D6E4F0"},
    SH_HVCB  : {"hdr": "7B3F00", "tab": "C55A11", "hdr2": "FDEBD0"},
    SH_BUS   : {"hdr": "1E5631", "tab": "548235", "hdr2": "D5F5E3"},
    SH_LEGEND: {"hdr": "4A235A", "tab": "8E44AD", "hdr2": "E8D5F5"},
}

# ── Section background fills ──────────────────────────────────────────────────
SECTION_FILLS = {
    "info"    : _fill("F0F3F4"),   # neutral grey
    "rated"   : _fill("FCE4D6"),   # warm orange
    "sim"     : _fill("DAEEF3"),   # cool blue
    "computed": _fill("E2EFDA"),   # green
}

# ── Overall result colour coding ──────────────────────────────────────────────
OVERALL_STYLE = {
    "Pass"           : (_fill("C6EFCE"), _font("276221", bold=True)),
    "Pass (Ik only)" : (_fill("D5F5E3"), _font("1E5631", bold=True)),
    "Pass (ip only)" : (_fill("D5F5E3"), _font("1E5631", bold=True)),
    "Fail"           : (_fill("FFC7CE"), _font("9C0006", bold=True)),
    "N/A – Insufficient Data": (_fill("F0F0F0"), _font("595959", bold=True)),
}

# ── Sub-result (I"k Result / ip Result) colour coding ────────────────────────
SUBRES_STYLE = {
    "Pass": (_fill("EBF5EB"), _font("276221")),
    "Fail": (_fill("FDE8EC"), _font("9C0006", bold=True)),
}


def _write_data_sheet(
    ws,
    df: pd.DataFrame,
    sheet_name: str,
) -> None:
    """Write one equipment group sheet with full formatting."""
    theme    = SHEET_THEME.get(sheet_name, SHEET_THEME[SH_LVCB])
    col_list = _ordered_cols(df)
    n_cols   = len(col_list)
    bdr      = _border()

    # ── Row 1: Main header (dark coloured) ────────────────────────────────────
    hdr_fill = _fill(theme["hdr"])
    for ci, col in enumerate(col_list, start=1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font      = _font(color="FFFFFF", bold=True, size=9)
        cell.fill      = hdr_fill
        cell.alignment = _align("center", wrap=True)
        cell.border    = bdr
    ws.row_dimensions[1].height = 30

    # ── Row 2: Section sub-header (section colour bands) ─────────────────────
    SECTION_LABELS = {
        "info"    : "",
        "rated"   : "◀  RATED  ▶",
        "sim"     : "◀  SIMULATED  ▶",
        "computed": "◀  EVALUATED  ▶",
    }
    hdr2_fill = _fill(theme["hdr2"])
    prev_section = None
    for ci, col in enumerate(col_list, start=1):
        section = _COL_SECTION.get(col, "info")
        label   = SECTION_LABELS.get(section, "")
        # Only write label at section boundary start
        cell = ws.cell(row=2, column=ci,
                       value=label if section != prev_section else "")
        cell.fill      = SECTION_FILLS.get(section, hdr2_fill)
        cell.font      = _font(color="444444", bold=True, italic=True, size=8)
        cell.alignment = _align("center")
        cell.border    = bdr
        prev_section   = section
    ws.row_dimensions[2].height = 13

    # ── Data rows (start at row 3) ────────────────────────────────────────────
    overall_ci = (col_list.index(C_OVERALL) + 1) if C_OVERALL in col_list else None
    res_ik_ci  = (col_list.index(C_RES_IK)  + 1) if C_RES_IK  in col_list else None
    res_ip_ci  = (col_list.index(C_RES_IP)  + 1) if C_RES_IP  in col_list else None
    util_ik_ci = (col_list.index(C_UTIL_IK) + 1) if C_UTIL_IK in col_list else None
    util_ip_ci = (col_list.index(C_UTIL_IP) + 1) if C_UTIL_IP in col_list else None

    ODD_FILL  = _fill("FFFFFF")
    EVEN_FILL = _fill("F5F9FD")

    for ri, (_, row) in enumerate(df[col_list].iterrows(), start=3):
        row_fill = ODD_FILL if ri % 2 == 1 else EVEN_FILL

        for ci, col in enumerate(col_list, start=1):
            val  = row[col]
            # Convert pandas NA to None
            if pd.isna(val) if not isinstance(val, str) else False:
                val = None
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = _font()
            cell.fill      = row_fill
            cell.alignment = _align()
            cell.border    = bdr

            # Number format for numeric simulation columns
            if isinstance(val, float) and col not in (
                C_OVERALL, C_RES_IK, C_RES_IP, C_REMARKS
            ):
                cell.number_format = "0.0000"
                cell.alignment = _align("right")

        # ── Special: Utilization % ────────────────────────────────────────────
        for uc in (util_ik_ci, util_ip_ci):
            if uc:
                c = ws.cell(row=ri, column=uc)
                if c.value is not None:
                    c.number_format = "0.00"
                    c.alignment     = _align("right")
                    # Traffic-light colour on utilization
                    try:
                        v = float(c.value)
                        if v > 100:
                            c.fill = _fill("FFC7CE")
                            c.font = _font("9C0006", bold=True)
                        elif v >= 80:
                            c.fill = _fill("FFEB9C")
                            c.font = _font("7D4E00")
                        else:
                            c.fill = _fill("EBF5EB")
                            c.font = _font("1E5631")
                    except (TypeError, ValueError):
                        pass

        # ── Special: Sub-results ──────────────────────────────────────────────
        for rc in (res_ik_ci, res_ip_ci):
            if rc:
                c = ws.cell(row=ri, column=rc)
                v = str(c.value or "").strip()
                if v in SUBRES_STYLE:
                    c.fill, c.font = SUBRES_STYLE[v]
                c.alignment = _align("center")

        # ── Special: Overall Result ───────────────────────────────────────────
        if overall_ci:
            c = ws.cell(row=ri, column=overall_ci)
            v = str(c.value or "").strip()
            if v in OVERALL_STYLE:
                c.fill, c.font = OVERALL_STYLE[v]
            c.alignment = _align("center")

    # ── Column widths ─────────────────────────────────────────────────────────
    FIXED_WIDTHS = {
        C_ID       : 26,
        C_BUS      : 24,
        C_OVERALL  : 22,
        C_RES_IK   : 14,
        C_RES_IP   : 14,
        C_REMARKS  : 22,
        C_UTIL_IK  : 18,
        C_UTIL_IP  : 18,
    }
    for ci, col in enumerate(col_list, start=1):
        letter = get_column_letter(ci)
        if col in FIXED_WIDTHS:
            ws.column_dimensions[letter].width = FIXED_WIDTHS[col]
        else:
            # Measure max content length
            max_len = max(
                (len(str(ws.cell(row=r, column=ci).value or ""))
                 for r in range(1, ws.max_row + 1)),
                default=8,
            )
            ws.column_dimensions[letter].width = min(max_len + 3, 22)

    # ── Freeze panes: rows 1+2, first column ─────────────────────────────────
    ws.freeze_panes = "B3"

    # ── Auto-filter ───────────────────────────────────────────────────────────
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(n_cols)}{ws.max_row}"
    )

    # ── Tab colour ────────────────────────────────────────────────────────────
    ws.sheet_properties.tabColor = theme["tab"]

    # ── Print settings ────────────────────────────────────────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1


def _write_legend_sheet(ws) -> None:
    """Write a colour-coded legend sheet explaining all columns and rules."""
    theme = SHEET_THEME[SH_LEGEND]
    ws.sheet_properties.tabColor = theme["tab"]
    bdr = _border()

    def _hdr(text: str, row: int) -> None:
        c = ws.cell(row=row, column=1, value=text)
        c.font      = _font("FFFFFF", bold=True, size=10)
        c.fill      = _fill(theme["hdr"])
        c.alignment = _align("left")
        c.border    = bdr
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row,   end_column=3
        )

    def _row(col_a: str, col_b: str, col_c: str,
             row: int, bg: str = "FFFFFF") -> None:
        for ci, val in enumerate([col_a, col_b, col_c], 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font      = _font(size=9)
            c.fill      = _fill(bg)
            c.alignment = _align("left")
            c.border    = bdr

    r = 1
    _hdr("⚡  SHORT CIRCUIT STUDY PROCESSOR — LEGEND", r); r += 1

    _hdr("COLUMN SECTIONS", r); r += 1
    _row("Section",         "Background",   "Description",             r, "E8EAF6"); r += 1
    _row("INFO",            "Light grey",   "Equipment ID, kV, Type, Bus",r, "F0F3F4"); r += 1
    _row("◀ RATED ▶",      "Warm orange",  "Nameplate rated values from equipment database",r,"FCE4D6"); r += 1
    _row("◀ SIMULATED ▶",  "Cool blue",    "Values from SC simulation software (I\"k, ip, Ib…)",r,"DAEEF3"); r += 1
    _row("◀ EVALUATED ▶",  "Green",        "Computed by this engine: Utilization %, Results",r,"E2EFDA"); r += 1

    r += 1
    _hdr("KEY COLUMN MAPPING", r); r += 1
    _row("Column",           "Maps To / Formula",            "Notes",    r, "E8EAF6"); r += 1
    _row("Rated Ib Sym",     "→ Rated Ik (kA)",              "Used as the breaking current rated duty reference (IEC 60909 rated duty Ib)",r); r += 1
    _row("I\"k",             "Simulated breaking current",   "Initial symmetrical SC current (kA) from simulation",r); r += 1
    _row("ip",               "Simulated peak current",       "Peak short-circuit current (kA) from simulation",r); r += 1
    _row("Rated ip",         "Rated peak current",           "Nameplate peak making/withstand current (kA)",r); r += 1
    _row("Rated Ik (kA)",    "= Rated Ib Sym",               "Explicit alias added for clarity",r); r += 1
    _row('Utilization % (I"k)',"= I\"k / Rated Ib Sym × 100","Breaking current loading percentage",r); r += 1
    _row("Utilization % (ip)", "= ip  / Rated ip  × 100",   "Peak current loading percentage",r); r += 1

    r += 1
    _hdr("EVALUATION RULES", r); r += 1
    _row("Check",            "Condition",                    "Result",   r, "E8EAF6"); r += 1
    _row("I\"k Check",       "I\"k (sim) ≤ Rated Ib Sym",   "Pass ✅",  r, "C6EFCE"); r += 1
    _row("I\"k Check",       "I\"k (sim) > Rated Ib Sym",   "Fail ❌",  r, "FFC7CE"); r += 1
    _row("ip Check",         "ip  (sim) ≤ Rated ip",        "Pass ✅",  r, "C6EFCE"); r += 1
    _row("ip Check",         "ip  (sim) > Rated ip",        "Fail ❌",  r, "FFC7CE"); r += 1
    _row("ip Check",         "Rated ip = 0",                "N/A (not rated for peak duty)",r,"FFFBEB"); r += 1

    r += 1
    _hdr("OVERALL RESULT COLOUR CODING", r); r += 1
    _row("Overall Result",   "Meaning",    "",  r, "E8EAF6"); r += 1
    _row("Pass",             "Both I\"k and ip checks passed",           "", r, "C6EFCE"); r += 1
    _row("Pass (Ik only)",   "I\"k passed; ip not applicable (Rated ip = 0)","", r, "D5F5E3"); r += 1
    _row("Fail",             "One or both checks exceeded rated duty",   "", r, "FFC7CE"); r += 1
    _row("N/A – Insuff.Data","Missing numeric data — manual review required","",r,"F0F0F0"); r += 1

    r += 1
    _hdr("UTILIZATION % COLOUR CODING", r); r += 1
    _row("0 – 79 %",         "Green  — within comfortable margin",  "", r, "EBF5EB"); r += 1
    _row("80 – 100 %",       "Amber  — approaching rated duty",     "", r, "FFEB9C"); r += 1
    _row("> 100 %",          "Red    — EXCEEDS rated duty → Fail",  "", r, "FFC7CE"); r += 1

    r += 1
    _hdr("SHEET CLASSIFICATION", r); r += 1
    _row("Sheet",    "Condition",              "Tab Colour",             r, "E8EAF6"); r += 1
    _row("LVCB",     "kV ≤ 1.0",              "Blue  — Low Voltage CB", r, "DAEEF3"); r += 1
    _row("HVCB",     "kV > 1.0",              "Orange — High Voltage CB",r,"FDEBD0"); r += 1
    _row("BUS",      "Type contains 'bus'",   "Green — Busbar / Busduct",r,"D5F5E3"); r += 1

    r += 1
    _hdr("STANDARD", r); r += 1
    _row("IEC 60909",
         "Short-circuit currents in three-phase a.c. systems",
         "Parts 0 (1988) / 1 (2016) / 4 (2000)",
         r); r += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 48
    ws.sheet_properties.tabColor    = theme["tab"]


def _write_summary_sheet(ws, results: dict[str, pd.DataFrame]) -> None:
    """Write a high-level summary sheet with counts per sheet."""
    bdr   = _border()
    hdr_f = _fill("1A1A2E")

    title = ws.cell(row=1, column=1, value="SHORT CIRCUIT STUDY — SUMMARY REPORT")
    title.font      = _font("FFFFFF", bold=True, size=12)
    title.fill      = hdr_f
    title.alignment = _align("center")
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 28

    # Sub-header
    headers = ["Sheet", "Total", "Pass", "Pass (Ik only)",
               "Fail", "N/A", "Pass Rate %", "Max Util% Ik"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = _font("FFFFFF", bold=True, size=9)
        c.fill = _fill("2E75B6")
        c.alignment = _align("center")
        c.border = bdr

    grand = {k: 0 for k in ["total", "pass", "xpass", "fail", "na"]}

    for ri, (sname, df) in enumerate(results.items(), start=3):
        total = len(df)
        vc    = df[C_OVERALL].value_counts()
        pass_  = int(vc.get("Pass", 0))
        xpass  = int(vc.get("Pass (Ik only)", 0)) + int(vc.get("Pass (ip only)", 0))
        fail_  = int(vc.get("Fail", 0))
        na_    = int(vc.get("N/A – Insufficient Data", 0))
        pr     = round((pass_ + xpass) / total * 100, 1) if total else 0
        max_uk = df[C_UTIL_IK].max() if C_UTIL_IK in df.columns else None

        row_fill = _fill("F5F9FD") if ri % 2 == 0 else _fill("FFFFFF")
        vals     = [sname, total, pass_, xpass, fail_, na_, pr,
                    round(max_uk, 1) if max_uk is not None else ""]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font      = _font(bold=(ci == 1))
            c.fill      = row_fill
            c.alignment = _align("center" if ci > 1 else "left")
            c.border    = bdr

        # Colour the fail count
        fc = ws.cell(row=ri, column=5)
        if fail_ > 0:
            fc.fill = _fill("FFC7CE")
            fc.font = _font("9C0006", bold=True)
        else:
            fc.fill = _fill("C6EFCE")
            fc.font = _font("276221")

        grand["total"] += total
        grand["pass"]  += pass_
        grand["xpass"] += xpass
        grand["fail"]  += fail_
        grand["na"]    += na_

    # Grand total row
    gr = len(results) + 3
    gt_pass  = grand["pass"] + grand["xpass"]
    gt_total = grand["total"]
    gt_pr    = round(gt_pass / gt_total * 100, 1) if gt_total else 0
    gt_vals  = ["TOTAL", gt_total, grand["pass"], grand["xpass"],
                grand["fail"], grand["na"], gt_pr, ""]
    for ci, v in enumerate(gt_vals, 1):
        c = ws.cell(row=gr, column=ci, value=v)
        c.font      = _font("FFFFFF", bold=True, size=10)
        c.fill      = _fill("1A1A2E")
        c.alignment = _align("center" if ci > 1 else "left")
        c.border    = bdr

    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 18

    ws.column_dimensions["A"].width = 22
    ws.sheet_properties.tabColor    = "1A1A2E"
    ws.freeze_panes = "A3"


def build_excel(results: dict[str, pd.DataFrame]) -> bytes:
    """
    Build the complete output Excel workbook in memory and return bytes.

    Sheet order: Summary | LVCB | HVCB | BUS | Legend
    """
    output = io.BytesIO()

    # Use openpyxl directly for full formatting control
    from openpyxl import Workbook
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # ── Summary ───────────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    _write_summary_sheet(ws_sum, results)

    # ── Data sheets ───────────────────────────────────────────────────────────
    sheet_order = [SH_LVCB, SH_HVCB, SH_BUS]
    for sname in sheet_order:
        if sname in results and not results[sname].empty:
            ws = wb.create_sheet(sname)
            _write_data_sheet(ws, results[sname], sname)

    # ── Legend ────────────────────────────────────────────────────────────────
    ws_leg = wb.create_sheet(SH_LEGEND)
    _write_legend_sheet(ws_leg)

    wb.save(output)
    return output.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
#  ❻  MAIN PROCESSING ORCHESTRATOR
# ══════════════════════════════════════════════════════════════════════════════

def process_upload(file_bytes: bytes) -> tuple[dict[str, pd.DataFrame], bytes, list[str]]:
    """
    Full pipeline: read → validate → evaluate → classify → build Excel.

    Returns
    -------
    results    : dict of sheet_name → evaluated DataFrame (for Streamlit preview)
    excel_bytes: final Excel file as bytes (for download button)
    warnings   : list of non-fatal warning messages
    """
    raw_df, warns = read_and_clean(file_bytes)
    evaluated     = evaluate_dataframe(raw_df)

    results: dict[str, pd.DataFrame] = {}
    for sheet_name in (SH_LVCB, SH_HVCB, SH_BUS):
        group = evaluated[evaluated["_sheet"] == sheet_name].copy()
        if not group.empty:
            results[sheet_name] = group

    if not results:
        raise ValueError("All rows were filtered out. Check that kV and Type columns have valid data.")

    excel_bytes = build_excel(results)
    return results, excel_bytes, warns


# ══════════════════════════════════════════════════════════════════════════════
#  ❼  STREAMLIT UI
# ══════════════════════════════════════════════════════════════════════════════

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="SC Study Processor",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Custom CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;700&family=IBM+Plex+Sans:wght@300;400;600;700&display=swap');

html, body, [class*="css"] {
    font-family: 'IBM Plex Sans', sans-serif;
}

/* ── Hero banner ─────────────────────────────────────────────────────── */
.hero {
    background: linear-gradient(135deg, #0A1628 0%, #1A3A5C 50%, #0D2137 100%);
    border: 1px solid #2E75B6;
    border-radius: 12px;
    padding: 2.2rem 2.5rem 2rem;
    margin-bottom: 1.5rem;
    position: relative;
    overflow: hidden;
}
.hero::before {
    content: '';
    position: absolute;
    top: -40px; right: -40px;
    width: 200px; height: 200px;
    background: radial-gradient(circle, #2E75B640 0%, transparent 70%);
}
.hero-title {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 1.7rem;
    font-weight: 700;
    color: #E8F4FD;
    margin: 0 0 .3rem;
    letter-spacing: -0.02em;
}
.hero-sub {
    color: #7FB3D3;
    font-size: .9rem;
    font-weight: 300;
    margin: 0;
    letter-spacing: 0.05em;
}
.hero-tags {
    margin-top: .9rem;
    display: flex;
    gap: .5rem;
    flex-wrap: wrap;
}
.tag {
    background: #2E75B620;
    border: 1px solid #2E75B660;
    color: #90C8E8;
    border-radius: 20px;
    padding: 2px 10px;
    font-size: .72rem;
    font-family: 'IBM Plex Mono', monospace;
    letter-spacing: 0.04em;
}

/* ── Info boxes ──────────────────────────────────────────────────────── */
.box {
    border-radius: 8px;
    padding: .85rem 1.1rem;
    margin: .4rem 0;
    font-size: .88rem;
    line-height: 1.55;
}
.box-info    { background:#EFF6FF; border-left:3px solid #2E75B6; color:#1a2e40; }
.box-success { background:#F0FDF4; border-left:3px solid #22C55E; color:#14532d; }
.box-warn    { background:#FFFBEB; border-left:3px solid #F59E0B; color:#78350f; }
.box-error   { background:#FEF2F2; border-left:3px solid #EF4444; color:#7f1d1d; }

/* ── Metric cards ────────────────────────────────────────────────────── */
.metric-card {
    background: #F8FAFC;
    border: 1px solid #E2E8F0;
    border-radius: 10px;
    padding: 1rem 1.2rem;
    text-align: center;
}
.metric-val  { font-size: 2rem; font-weight: 700; line-height: 1; }
.metric-lbl  { font-size: .75rem; color: #64748B; margin-top: .3rem; text-transform: uppercase; letter-spacing: .06em; }
.metric-pass { color: #16A34A; }
.metric-fail { color: #DC2626; }
.metric-warn { color: #D97706; }
.metric-neu  { color: #1F4E79; }

/* ── Section badge pills ─────────────────────────────────────────────── */
.pill-rated    { background:#FCE4D6; color:#7B3F00; border-radius:12px; padding:1px 9px; font-size:.75rem; font-weight:600; }
.pill-sim      { background:#DAEEF3; color:#1F4E79; border-radius:12px; padding:1px 9px; font-size:.75rem; font-weight:600; }
.pill-computed { background:#E2EFDA; color:#1E5631; border-radius:12px; padding:1px 9px; font-size:.75rem; font-weight:600; }

/* ── Download button ─────────────────────────────────────────────────── */
.stDownloadButton > button {
    background: linear-gradient(135deg, #1F4E79, #2E75B6) !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    font-size: 1rem !important;
    padding: .7rem 1.5rem !important;
    letter-spacing: .02em !important;
}
.stDownloadButton > button:hover {
    background: linear-gradient(135deg, #2E75B6, #1F4E79) !important;
    box-shadow: 0 4px 12px rgba(30,78,121,.3) !important;
}

footer {visibility: hidden;}
#MainMenu {visibility: hidden;}
.stDeployButton {display: none;}
</style>
""", unsafe_allow_html=True)


# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style='text-align:center; padding:.5rem 0 1rem'>
      <div style='font-size:2.5rem'>⚡</div>
      <div style='font-family:"IBM Plex Mono",monospace; font-weight:700;
                  color:#1F4E79; font-size:1rem; letter-spacing:-.01em'>
        SC STUDY PROCESSOR
      </div>
      <div style='color:#64748B; font-size:.75rem; margin-top:.2rem'>
        IEC 60909 · Phase 1
      </div>
    </div>
    """, unsafe_allow_html=True)

    st.divider()

    with st.expander("📋 Required Input Columns", expanded=True):
        for c in REQUIRED_COLS:
            st.markdown(f"`{c}`")

    with st.expander("🔢 Column Mapping", expanded=True):
        st.markdown("""
<span class="pill-rated">RATED</span> from nameplate data:
- `Rated ip` → Rated peak current (kA)
- `Rated Ib Sym` → **Rated Ik** (breaking)

<span class="pill-sim">SIMULATED</span> from SC software:
- `I"k` → Calculated breaking current
- `ip` → Calculated peak current

<span class="pill-computed">EVALUATED</span> computed by engine:
- `Rated Ik (kA)` = Rated Ib Sym
- `Utilization % (I"k)` = I"k÷RatedIbSym×100
- `Utilization % (ip)` = ip÷Ratedip×100
- `I"k Result` → Pass / Fail
- `ip Result` → Pass / Fail / N/A
- `Overall Result` → combined verdict
""", unsafe_allow_html=True)

    with st.expander("⚖️ Evaluation Logic", expanded=False):
        st.markdown("""
**CHECK 1 — Breaking (I"k)**
```
I"k ≤ Rated Ib Sym  →  Pass
I"k > Rated Ib Sym  →  Fail
```
**CHECK 2 — Peak (ip)**
```
ip ≤ Rated ip       →  Pass
ip > Rated ip       →  Fail
Rated ip = 0        →  N/A
```
**Overall**
```
Both Pass    →  Pass
Any Fail     →  Fail
Ik only      →  Pass (Ik only)
```
""")

    with st.expander("🗂️ Sheet Classification", expanded=False):
        st.markdown("""
| Condition | Sheet |
|-----------|-------|
| kV ≤ 1.0 | 🔵 **LVCB** |
| kV > 1.0 | 🟠 **HVCB** |
| Type = bus | 🟢 **BUS** |
""")


# ── Hero banner ───────────────────────────────────────────────────────────────
st.markdown("""
<div class="hero">
  <div class="hero-title">⚡ Short Circuit Study Processor</div>
  <div class="hero-sub">AUTOMATED IEC 60909 EVALUATION ENGINE · DUAL-CHECK · AUTO-CLASSIFICATION</div>
  <div class="hero-tags">
    <span class="tag">I"k Breaking Check</span>
    <span class="tag">ip Peak Check</span>
    <span class="tag">Utilization %</span>
    <span class="tag">LVCB / HVCB / BUS Split</span>
    <span class="tag">Rated Ib Sym → Rated Ik Mapping</span>
    <span class="tag">Colour-coded Excel Output</span>
  </div>
</div>
""", unsafe_allow_html=True)


# ── Upload section ────────────────────────────────────────────────────────────
col_up, col_guide = st.columns([3, 2], gap="large")

with col_up:
    uploaded = st.file_uploader(
        "Upload SC Result Excel (.xlsx)",
        type=["xlsx"],
        label_visibility="visible",
        help="Supports single or multi-sheet workbooks. All sheets are scanned for valid equipment data.",
    )

with col_guide:
    st.markdown("""
<div class="box box-info">
<strong>📌 What this engine does</strong><br><br>
1. Reads every sheet in your workbook<br>
2. Strips header whitespace, validates columns<br>
3. Maps <code>Rated Ib Sym</code> → <code>Rated Ik (kA)</code><br>
4. Runs <strong>two independent checks</strong> per row:<br>
&nbsp;&nbsp;&nbsp;• Breaking: I"k vs Rated Ib Sym<br>
&nbsp;&nbsp;&nbsp;• Peak: ip vs Rated ip<br>
5. Computes Utilization % for each<br>
6. Auto-classifies into LVCB / HVCB / BUS<br>
7. Produces formatted Excel with Legend<br>
8. <strong>One-click download ↓</strong>
</div>
""", unsafe_allow_html=True)


# ── Processing ────────────────────────────────────────────────────────────────
if uploaded is not None:
    st.divider()

    # File info row
    kb = round(uploaded.size / 1024, 1)
    st.markdown(
        f'<div class="box box-info">📄 <strong>{uploaded.name}</strong>'
        f' &nbsp;|&nbsp; {kb} KB &nbsp;|&nbsp; {uploaded.type or "application/xlsx"}</div>',
        unsafe_allow_html=True,
    )

    if st.button("⚙️  Run SC Evaluation", type="primary", use_container_width=True):

        with st.spinner("Processing… validating → evaluating → building Excel…"):
            try:
                results, excel_bytes, warns = process_upload(uploaded.getvalue())
                st.session_state["results"]     = results
                st.session_state["excel_bytes"] = excel_bytes
                st.session_state["warns"]       = warns
                st.session_state["filename"]    = uploaded.name
            except ValueError as e:
                st.markdown(
                    f'<div class="box box-error">❌ <strong>Error:</strong> {e}</div>',
                    unsafe_allow_html=True,
                )
                st.stop()
            except Exception as e:
                st.markdown(
                    f'<div class="box box-error">❌ <strong>Unexpected error:</strong> {e}</div>',
                    unsafe_allow_html=True,
                )
                st.stop()

    # Show results if available
    if "results" in st.session_state:
        results     = st.session_state["results"]
        excel_bytes = st.session_state["excel_bytes"]
        warns       = st.session_state["warns"]
        orig_name   = st.session_state["filename"]

        # Non-fatal warnings
        for w in warns:
            st.markdown(
                f'<div class="box box-warn">⚠️ {w}</div>',
                unsafe_allow_html=True,
            )

        st.markdown('<div class="box box-success">✅ <strong>Evaluation complete!</strong></div>',
                    unsafe_allow_html=True)

        # ────────────────────────────────────────────────────────────────────
        # SUMMARY METRICS
        # ────────────────────────────────────────────────────────────────────
        st.markdown("### 📊 Summary")

        all_rows = pd.concat(results.values(), ignore_index=True)
        total    = len(all_rows)
        vc       = all_rows[C_OVERALL].value_counts()
        n_pass   = int(vc.get("Pass", 0))
        n_xpass  = int(vc.get("Pass (Ik only)", 0)) + int(vc.get("Pass (ip only)", 0))
        n_fail   = int(vc.get("Fail", 0))
        n_na     = int(vc.get("N/A – Insufficient Data", 0))
        pass_rate = round((n_pass + n_xpass) / total * 100, 1) if total else 0
        max_util  = all_rows[C_UTIL_IK].max() if C_UTIL_IK in all_rows else None

        m1, m2, m3, m4, m5, m6 = st.columns(6)
        with m1:
            st.markdown(f"""<div class="metric-card">
              <div class="metric-val metric-neu">{total}</div>
              <div class="metric-lbl">Total Equipment</div></div>""",
              unsafe_allow_html=True)
        with m2:
            st.markdown(f"""<div class="metric-card">
              <div class="metric-val metric-pass">{n_pass}</div>
              <div class="metric-lbl">Pass</div></div>""",
              unsafe_allow_html=True)
        with m3:
            st.markdown(f"""<div class="metric-card">
              <div class="metric-val metric-warn">{n_xpass}</div>
              <div class="metric-lbl">Pass (Ik only)</div></div>""",
              unsafe_allow_html=True)
        with m4:
            st.markdown(f"""<div class="metric-card">
              <div class="metric-val metric-fail">{n_fail}</div>
              <div class="metric-lbl">Fail</div></div>""",
              unsafe_allow_html=True)
        with m5:
            st.markdown(f"""<div class="metric-card">
              <div class="metric-val metric-pass" style="font-size:1.5rem">{pass_rate}%</div>
              <div class="metric-lbl">Pass Rate</div></div>""",
              unsafe_allow_html=True)
        with m6:
            max_str = f"{max_util:.1f}%" if max_util is not None else "—"
            color   = "metric-fail" if (max_util or 0) > 100 else "metric-warn"
            st.markdown(f"""<div class="metric-card">
              <div class="metric-val {color}" style="font-size:1.5rem">{max_str}</div>
              <div class="metric-lbl">Max Util% (I"k)</div></div>""",
              unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # ────────────────────────────────────────────────────────────────────
        # PER-SHEET BREAKDOWN
        # ────────────────────────────────────────────────────────────────────
        if len(results) > 1:
            st.markdown("### 🗂️ By Sheet")
            bcols = st.columns(len(results))
            icons = {SH_LVCB: "🔵", SH_HVCB: "🟠", SH_BUS: "🟢"}
            for col, (sname, sdf) in zip(bcols, results.items()):
                with col:
                    svc = sdf[C_OVERALL].value_counts()
                    sp  = int(svc.get("Pass", 0)) + int(svc.get("Pass (Ik only)", 0))
                    sf  = int(svc.get("Fail", 0))
                    st.markdown(f"""<div class="metric-card">
                      <div style="font-size:1.1rem;font-weight:700;color:#1F4E79">
                        {icons.get(sname,'⚪')} {sname}</div>
                      <div style="margin:.5rem 0;font-size:.85rem">
                        <span style="color:#16A34A;font-weight:700">{sp} Pass</span>
                        &nbsp;|&nbsp;
                        <span style="color:#DC2626;font-weight:700">{sf} Fail</span>
                      </div>
                      <div style="color:#64748B;font-size:.75rem">{len(sdf)} total</div>
                    </div>""", unsafe_allow_html=True)

        # ────────────────────────────────────────────────────────────────────
        # INTERACTIVE RESULTS TABLE
        # ────────────────────────────────────────────────────────────────────
        st.markdown("### 📋 Equipment Results")

        # Sheet selector tabs
        sheet_icons = {SH_LVCB: "🔵 LVCB", SH_HVCB: "🟠 HVCB", SH_BUS: "🟢 BUS"}
        tab_labels  = [sheet_icons.get(n, n) for n in results]
        tabs        = st.tabs(tab_labels)

        for tab, (sname, sdf) in zip(tabs, results.items()):
            with tab:
                # ── Filter row ────────────────────────────────────────────
                fc1, fc2, fc3 = st.columns([2, 2, 3])

                with fc1:
                    res_opts = ["All Results"] + sorted(sdf[C_OVERALL].unique().tolist())
                    res_filt = st.selectbox(
                        "Filter Overall Result",
                        res_opts,
                        key=f"rfilt_{sname}",
                    )
                with fc2:
                    ik_opts = ["All I\"k Results"] + sorted(sdf[C_RES_IK].unique().tolist())
                    ik_filt = st.selectbox(
                        'Filter I"k Result',
                        ik_opts,
                        key=f"ikfilt_{sname}",
                    )
                with fc3:
                    bus_opts = ["All Buses"] + sorted(sdf[C_BUS].dropna().unique().tolist())
                    bus_filt = st.selectbox(
                        "Filter by Bus",
                        bus_opts,
                        key=f"busfilt_{sname}",
                    )

                disp = sdf.copy()
                if res_filt != "All Results":
                    disp = disp[disp[C_OVERALL] == res_filt]
                if ik_filt != 'All I"k Results':
                    disp = disp[disp[C_RES_IK] == ik_filt]
                if bus_filt != "All Buses":
                    disp = disp[disp[C_BUS] == bus_filt]

                # ── Column picker ─────────────────────────────────────────
                display_cols_default = [
                    C_ID, C_BUS, C_RATED_IB_SYM, C_RATED_IK,
                    C_RATED_IP, C_IK, C_IP,
                    C_UTIL_IK, C_UTIL_IP,
                    C_RES_IK, C_RES_IP, C_OVERALL,
                ]
                avail = [c for c in _ordered_cols(disp) if c in disp.columns]
                defaults = [c for c in display_cols_default if c in avail]

                show_cols = st.multiselect(
                    "Columns to display",
                    options=avail,
                    default=defaults,
                    key=f"cols_{sname}",
                )
                if not show_cols:
                    show_cols = defaults

                disp = disp[[c for c in show_cols if c in disp.columns]]

                # ── Styling function ──────────────────────────────────────
                def _style_df(df_s: pd.DataFrame) -> pd.DataFrame:
                    """Apply cell-level background colours using pd.Styler."""
                    styled = df_s.style

                    def _bg_overall(val):
                        m = {
                            "Pass":           "background-color:#C6EFCE;color:#276221;font-weight:700",
                            "Pass (Ik only)": "background-color:#D5F5E3;color:#1E5631;font-weight:700",
                            "Pass (ip only)": "background-color:#D5F5E3;color:#1E5631;font-weight:700",
                            "Fail":           "background-color:#FFC7CE;color:#9C0006;font-weight:700",
                        }
                        return m.get(str(val).strip(), "")

                    def _bg_subres(val):
                        v = str(val).strip()
                        if v == "Pass": return "color:#276221"
                        if v == "Fail": return "background-color:#FDE8EC;color:#9C0006;font-weight:700"
                        return "color:#888888;font-style:italic"

                    def _bg_util(val):
                        try:
                            v = float(val)  # type: ignore[arg-type]
                            if v > 100: return "background-color:#FFC7CE;color:#9C0006;font-weight:700"
                            if v >= 80:  return "background-color:#FFEB9C;color:#7D4E00"
                            return "background-color:#EBF5EB;color:#1E5631"
                        except (TypeError, ValueError):
                            return ""

                    if C_OVERALL in df_s.columns:
                        styled = styled.applymap(_bg_overall, subset=[C_OVERALL])
                    for rc in (C_RES_IK, C_RES_IP):
                        if rc in df_s.columns:
                            styled = styled.applymap(_bg_subres, subset=[rc])
                    for uc in (C_UTIL_IK, C_UTIL_IP):
                        if uc in df_s.columns:
                            styled = styled.applymap(_bg_util, subset=[uc])

                    num_fmt = {
                        c: "{:.4f}" for c in df_s.columns
                        if df_s[c].dtype == "float64"
                        and c not in (C_UTIL_IK, C_UTIL_IP)
                    }
                    util_fmt = {
                        c: "{:.2f}" for c in (C_UTIL_IK, C_UTIL_IP)
                        if c in df_s.columns
                    }
                    if {**num_fmt, **util_fmt}:
                        styled = styled.format({**num_fmt, **util_fmt}, na_rep="—")

                    return styled

                st.dataframe(
                    _style_df(disp.head(500)),
                    use_container_width=True,
                    height=460,
                )

                if len(disp) > 500:
                    st.caption(f"Showing 500 of {len(disp)} rows. Full data in the download.")

                # ── Per-tab mini-stats ────────────────────────────────────
                if C_UTIL_IK in sdf.columns:
                    s1, s2, s3, s4 = st.columns(4)
                    vc2 = sdf[C_OVERALL].value_counts()
                    s1.metric("Pass",  int(vc2.get("Pass", 0)))
                    s2.metric("Fail",  int(vc2.get("Fail", 0)))
                    s3.metric('Max Util% I"k',
                              f"{sdf[C_UTIL_IK].max():.1f}%" if sdf[C_UTIL_IK].notna().any() else "—")
                    s4.metric("Max Util% ip",
                              f"{sdf[C_UTIL_IP].max():.1f}%" if C_UTIL_IP in sdf.columns and sdf[C_UTIL_IP].notna().any() else "—")

        # ────────────────────────────────────────────────────────────────────
        # FAIL SPOTLIGHT — quick view of all failures
        # ────────────────────────────────────────────────────────────────────
        all_fails = all_rows[all_rows[C_OVERALL] == "Fail"]
        if not all_fails.empty:
            with st.expander(f"🚨  View All {len(all_fails)} FAIL Items", expanded=False):
                fail_cols = [
                    C_ID, C_KV, C_BUS,
                    C_RATED_IB_SYM, C_IK, C_UTIL_IK, C_RES_IK,
                    C_RATED_IP,     C_IP, C_UTIL_IP,  C_RES_IP,
                    C_OVERALL,
                ]
                fail_cols = [c for c in fail_cols if c in all_fails.columns]
                fail_disp = all_fails[fail_cols].sort_values(C_UTIL_IK, ascending=False)
                st.dataframe(
                    fail_disp.style
                    .applymap(
                        lambda v: "background-color:#FFC7CE;color:#9C0006;font-weight:700"
                        if str(v) == "Fail" else "",
                        subset=[c for c in (C_RES_IK, C_RES_IP, C_OVERALL) if c in fail_disp.columns]
                    )
                    .format(
                        {c: "{:.2f}" for c in (C_UTIL_IK, C_UTIL_IP) if c in fail_disp.columns},
                        na_rep="—"
                    ),
                    use_container_width=True,
                    height=380,
                )

        # ────────────────────────────────────────────────────────────────────
        # DOWNLOAD
        # ────────────────────────────────────────────────────────────────────
        st.divider()
        st.markdown("### ⬇️ Download Processed Excel")

        dl_name = (
            uploaded.name.replace(".xlsx", "") + "_SC_Evaluated.xlsx"
            if uploaded else "SC_Processed.xlsx"
        )

        col_dl, col_hint = st.columns([2, 3])
        with col_dl:
            st.download_button(
                label="⬇️  Download Evaluated Excel",
                data=excel_bytes,
                file_name=dl_name,
                mime=(
                    "application/vnd.openxmlformats-officedocument"
                    ".spreadsheetml.sheet"
                ),
                use_container_width=True,
            )
        with col_hint:
            st.markdown("""
<div class="box box-info">
Output workbook contains:<br>
📊 <strong>Summary</strong> sheet — counts per sheet<br>
🔵 <strong>LVCB</strong> / 🟠 <strong>HVCB</strong> / 🟢 <strong>BUS</strong> sheets — full evaluated data<br>
🟣 <strong>Legend</strong> sheet — column guide & colour key<br>
All sheets: colour-coded results, auto-filter, freeze panes, print-ready
</div>""", unsafe_allow_html=True)

else:
    # ── Landing state: sample format guide ────────────────────────────────────
    st.markdown("### 📌 Expected Input Format")
    sample_df = pd.DataFrame({
        "ID":           ["CB5", "FDR-2 BUS OG", "HV-CB-01"],
        "kV":           [0.415,  0.415,           11.0],
        "Type":         ["CB",   "CB",            "CB"],
        "Cfactor":      [1.05,   1.05,            1.05],
        "Bus":          ["L2-UR01-COB", "FDR-2 BUS", "HV-BUS-01"],
        "Rated ip":     [73.5,   30.0,            52.5],
        "Rated Ib Sym": [35.0,   7.5,             25.0],
        'I"k':          [23.34,  22.33,           15.0],
        "ip":           [34.26,  34.97,           22.0],
    })
    st.dataframe(sample_df, use_container_width=True)

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("""
<div class="box box-info">
<strong>Sheet classification rules:</strong><br>
• kV ≤ 1.0 → <strong>LVCB</strong> sheet<br>
• kV &gt; 1.0 → <strong>HVCB</strong> sheet<br>
• Type contains "bus" → <strong>BUS</strong> sheet<br>
• Multiple input sheets supported (all scanned)
</div>""", unsafe_allow_html=True)
    with c2:
        st.markdown("""
<div class="box box-info">
<strong>Key mapping:</strong><br>
• <code>Rated Ib Sym</code> → used as <strong>Rated Ik</strong> (breaking duty reference)<br>
• <code>I"k</code> → simulated breaking current<br>
• <code>ip</code> → simulated peak current<br>
• <code>Rated ip</code> = 0 → peak check marked N/A
</div>""", unsafe_allow_html=True)
